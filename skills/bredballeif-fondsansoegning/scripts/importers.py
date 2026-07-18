"""Offline-first importers for the Bredballe IF fund index.

Only the workbook reader depends on ``openpyxl``. It is imported inside the
XLSX functions so JSON and CSV workflows continue to work without that
optional dependency installed.
"""

from __future__ import annotations

import csv
import io
import re
import tempfile
import unicodedata
from collections.abc import Iterable, Iterator, Mapping, Sequence
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, BinaryIO
from urllib.error import HTTPError, URLError
from urllib.parse import urlsplit, urlunsplit
from urllib.request import HTTPRedirectHandler, Request, build_opener
from zipfile import BadZipFile, ZipFile

try:  # Works both as a package and with scripts/ on PYTHONPATH.
    from .json_store import JsonFundStore as IndexStore, normalize_domain, normalize_verification
except ImportError:  # pragma: no cover - exercised by CLI-style imports.
    from json_store import JsonFundStore as IndexStore, normalize_domain, normalize_verification


KNOWN_FUND_SHEETS = ("02_Aktuelle", "03_Fondsindeks")
FUND_HEADER_ROW = 4
DGI_FUND_SHEET = "Fonde - oversigt"
DEFAULT_MAX_HISTORY_DOWNLOAD_BYTES = 25 * 1024 * 1024
MAX_XLSX_MEMBERS = 10_000
MAX_XLSX_UNCOMPRESSED_BYTES = 256 * 1024 * 1024
MAX_XLSX_MEMBER_BYTES = 64 * 1024 * 1024
MAX_SUSPICIOUS_COMPRESSION_RATIO = 1_000


class ImportFormatError(ValueError):
    """Raised when a supplied file does not contain the expected table."""


class ImportDependencyError(RuntimeError):
    """Raised when XLSX support was requested without openpyxl installed."""


class HistoryDownloadError(RuntimeError):
    """Raised when a shared history workbook cannot be downloaded safely."""


@dataclass
class ImportResult:
    """Machine-readable import summary used by the CLI and tests."""

    source: str
    records_seen: int = 0
    imported: int = 0
    inserted: int = 0
    updated: int = 0
    duplicates: int = 0
    skipped: int = 0
    errors: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors

    def to_dict(self) -> dict[str, Any]:
        result = asdict(self)
        result["ok"] = self.ok
        return result


@dataclass(frozen=True)
class _FundRow:
    record: dict[str, Any]
    raw: dict[str, Any]
    sheet: str
    row_number: int
    source_record_id: str


_HEADER_TRANSLATION = str.maketrans(
    {
        "æ": "ae",
        "ø": "oe",
        "å": "aa",
        "Æ": "ae",
        "Ø": "oe",
        "Å": "aa",
    }
)


def normalize_header(value: Any) -> str:
    """Normalize Danish/English column headers for alias matching."""

    if value is None:
        return ""
    text = str(value).strip().translate(_HEADER_TRANSLATION).casefold()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(character for character in text if not unicodedata.combining(character))
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def normalize_date_value(value: Any) -> str:
    """Normalize spreadsheet dates to ISO where possible, otherwise preserve text."""

    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # Excel's default 1900 date system, including its historical leap-year
        # compatibility offset, is represented by a 1899-12-30 epoch.
        if 1 <= float(value) < 100_000:
            converted = datetime(1899, 12, 30) + timedelta(days=float(value))
            if converted.time().isoformat() == "00:00:00":
                return converted.date().isoformat()
            return converted.isoformat(timespec="seconds")
        return str(value)

    text = str(value).strip()
    for pattern in (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%d.%m.%Y",
        "%Y-%m-%d %H:%M:%S",
        "%d-%m-%Y %H:%M:%S",
    ):
        try:
            parsed = datetime.strptime(text, pattern)
        except ValueError:
            continue
        return parsed.date().isoformat() if parsed.time() == datetime.min.time() else parsed.isoformat(timespec="seconds")
    return text


def _load_workbook(path: str | Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exact message tested indirectly.
        raise ImportDependencyError(
            "XLSX-import kræver openpyxl. Installér skillens requirements først."
        ) from exc
    candidate = Path(path)
    if candidate.is_file():
        _validate_xlsx_archive(candidate)
    return load_workbook(filename=path, read_only=True, data_only=True)


def _validate_xlsx_archive(path: str | Path) -> None:
    """Bound XLSX expansion before openpyxl parses attacker-controlled XML."""

    try:
        with ZipFile(path) as archive:
            members = archive.infolist()
            names = {item.filename for item in members}
            if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
                raise ImportFormatError("ZIP-filen er ikke en gyldig XLSX-projektfil")
            if len(members) > MAX_XLSX_MEMBERS:
                raise ImportFormatError("XLSX-filen indeholder for mange ZIP-medlemmer")
            total = 0
            for item in members:
                if item.flag_bits & 0x1:
                    raise ImportFormatError("Krypterede ZIP-medlemmer understøttes ikke")
                total += int(item.file_size)
                if item.file_size > MAX_XLSX_MEMBER_BYTES:
                    raise ImportFormatError("Et XLSX-medlem overskrider størrelsesgrænsen")
                ratio = item.file_size / max(1, item.compress_size)
                if item.file_size > 10 * 1024 * 1024 and ratio > MAX_SUSPICIOUS_COMPRESSION_RATIO:
                    raise ImportFormatError("XLSX-filen har en mistænkelig kompressionsratio")
            if total > MAX_XLSX_UNCOMPRESSED_BYTES:
                raise ImportFormatError("XLSX-filens ukomprimerede størrelse overskrider grænsen")
    except BadZipFile as exc:
        raise ImportFormatError("Filen er ikke en gyldig XLSX/ZIP-fil") from exc


_FUND_SHEET_FIELDS: dict[str, dict[str, str]] = {
    "02_Aktuelle": {
        "id": "source_record_id",
        "type": "type",
        "fond pulje portal": "name",
        "geografi": "geography",
        "hvem kan soege": "applicant_types",
        "typisk formaal": "purposes",
        "beloeb ramme": "amount",
        "frist frekvens": "deadline",
        "vigtige krav": "requirements",
        "vigtige udelukkelser": "exclusions",
        "status": "verification_status",
        "sidst kontrolleret": "last_checked",
        "officiel url": "url",
        "praktisk note": "notes",
    },
    "03_Fondsindeks": {
        "indeks id": "source_record_id",
        "geografi": "geography",
        "kommune omraade": "area",
        "fond pulje": "name",
        "kort beskrivelse": "description",
        "url": "url",
        "kilde": "listed_source",
        "verifikationsstatus": "verification_status",
        "sidst kontrolleret": "last_checked",
        "relevans for bredballe if": "relevance",
        "noter": "notes",
    },
}


def _raw_row(headers: Sequence[Any], values: Sequence[Any]) -> dict[str, Any]:
    raw: dict[str, Any] = {}
    for index, header in enumerate(headers):
        if header in (None, ""):
            continue
        key = str(header).strip()
        if key in raw:
            key = f"{key} ({index + 1})"
        raw[key] = values[index] if index < len(values) else None
    return raw


def _iter_fund_rows(
    workbook_path: str | Path,
    sheet_names: Sequence[str] = KNOWN_FUND_SHEETS,
) -> Iterator[_FundRow]:
    workbook = _load_workbook(workbook_path)
    found_sheet = False
    try:
        for sheet_name in sheet_names:
            if sheet_name not in workbook.sheetnames:
                continue
            found_sheet = True
            sheet = workbook[sheet_name]
            rows = iter(sheet.iter_rows(min_row=FUND_HEADER_ROW, values_only=True))
            try:
                headers = tuple(next(rows))
            except StopIteration:
                continue

            field_map = _FUND_SHEET_FIELDS.get(sheet_name, {})
            mapped_headers = [field_map.get(normalize_header(header), "") for header in headers]
            if "name" not in mapped_headers:
                raise ImportFormatError(
                    f"{sheet_name}: række {FUND_HEADER_ROW} mangler fondskolonnen"
                )

            for row_number, values in enumerate(rows, start=FUND_HEADER_ROW + 1):
                values = tuple(values)
                if not any(value not in (None, "") for value in values):
                    continue
                raw = _raw_row(headers, values)
                record: dict[str, Any] = {}
                listed_source = ""
                source_record_id = ""
                for index, canonical_field in enumerate(mapped_headers):
                    if not canonical_field or index >= len(values):
                        continue
                    value = values[index]
                    if value in (None, ""):
                        continue
                    if canonical_field == "source_record_id":
                        source_record_id = str(value).strip()
                    elif canonical_field == "listed_source":
                        listed_source = str(value).strip()
                    elif canonical_field == "last_checked":
                        record[canonical_field] = normalize_date_value(value)
                    elif canonical_field == "verification_status":
                        record[canonical_field] = normalize_verification(value)
                    else:
                        record[canonical_field] = value

                if not str(record.get("name", "")).strip():
                    # A non-empty formatting row is not an import error.
                    continue
                if listed_source:
                    record["extra"] = {"listed_source": listed_source}
                yield _FundRow(
                    record=record,
                    raw=raw,
                    sheet=sheet_name,
                    row_number=row_number,
                    source_record_id=source_record_id or f"row-{row_number}",
                )
    finally:
        workbook.close()
    if not found_sheet:
        names = ", ".join(sheet_names)
        raise ImportFormatError(f"Ingen kendte fondsark fundet ({names})")


def read_fund_workbook(
    workbook_path: str | Path,
    *,
    sheet_names: Sequence[str] = KNOWN_FUND_SHEETS,
) -> list[dict[str, Any]]:
    """Read canonical fund records without modifying the JSON store."""

    return [row.record for row in _iter_fund_rows(workbook_path, sheet_names)]


def import_fund_workbook(
    workbook_path: str | Path,
    store: IndexStore,
    *,
    source_name: str | None = None,
    source_url: str | None = None,
    sheet_names: Sequence[str] = KNOWN_FUND_SHEETS,
) -> ImportResult:
    """Import the known ``02_Aktuelle`` and ``03_Fondsindeks`` tables."""

    path = Path(workbook_path)
    safe_source_name = source_name or path.name
    result = ImportResult(source=safe_source_name)
    known_ids = {int(item["fund_id"]) for item in store.list_funds()}
    for row in _iter_fund_rows(path, sheet_names):
        result.records_seen += 1
        try:
            fund_id = store.upsert_fund(
                row.record,
                source_name=f"{safe_source_name}#{row.sheet}",
                source_record_id=row.source_record_id,
                source_url=source_url,
                source_kind="xlsx",
                raw=row.raw,
            )
        except (TypeError, ValueError) as exc:
            result.skipped += 1
            result.errors.append(f"{row.sheet}, række {row.row_number}: {exc}")
            continue
        result.imported += 1
        if fund_id not in known_ids:
            result.inserted += 1
            known_ids.add(fund_id)
        else:
            result.updated += 1
    return result


def import_dgi_workbook(
    workbook_path: str | Path,
    store: IndexStore,
    *,
    source_name: str = "DGI – aktuel fondsliste",
    source_url: str | None = None,
) -> ImportResult:
    """Import DGI's public four-column fund workbook as discovery records."""

    workbook = _load_workbook(workbook_path)
    result = ImportResult(source=source_name)
    known_ids = {int(item["fund_id"]) for item in store.list_funds()}
    try:
        if DGI_FUND_SHEET not in workbook.sheetnames:
            raise ImportFormatError(f"DGI-arket mangler: {DGI_FUND_SHEET}")
        sheet = workbook[DGI_FUND_SHEET]
        rows = list(sheet.iter_rows(values_only=True))
        header_index = next(
            (
                index
                for index, values in enumerate(rows[:30])
                if {normalize_header(value) for value in values if value not in (None, "")}
                >= {"kommune", "fond", "kort beskrivelse", "link"}
            ),
            None,
        )
        if header_index is None:
            raise ImportFormatError("DGI-arket mangler kolonnerne Kommune, Fond, Kort beskrivelse og Link")
        headers = tuple(rows[header_index])
        positions = {normalize_header(value): index for index, value in enumerate(headers)}
        for row_number, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
            area = str(values[positions["kommune"]] or "").strip()
            name = str(values[positions["fond"]] or "").strip()
            description = str(values[positions["kort beskrivelse"]] or "").strip()
            url = str(values[positions["link"]] or "").strip()
            if not name:
                continue
            result.records_seen += 1
            record = {
                "name": name,
                "url": url,
                "type": "fond eller pulje",
                "geography": "Danmark" if normalize_header(area) == "nationalt" else area,
                "area": area,
                "description": description,
                "verification_status": "unverified",
                "extra": {"listed_source": "DGI's offentlige fondsoversigt"},
            }
            raw = _raw_row(headers, values)
            try:
                fund_id = store.upsert_fund(
                    record,
                    source_name=f"{source_name}#{DGI_FUND_SHEET}",
                    source_record_id=f"row-{row_number}",
                    source_url=source_url,
                    source_kind="official_directory",
                    raw=raw,
                )
            except (TypeError, ValueError) as exc:
                result.skipped += 1
                result.errors.append(f"{DGI_FUND_SHEET}, række {row_number}: {exc}")
                continue
            result.imported += 1
            if fund_id not in known_ids:
                result.inserted += 1
                known_ids.add(fund_id)
            else:
                result.updated += 1
    finally:
        workbook.close()
    return result


_HISTORY_ALIASES: dict[str, set[str]] = {
    "fund_name": {
        "fond",
        "fondsnavn",
        "fond pulje",
        "fond pulje portal",
        "pulje",
        "stoettegiver",
        "udbyder",
        "fund",
        "funder",
        "fund name",
        "funding source",
    },
    "fund_url": {
        "url",
        "link",
        "hjemmeside",
        "fondsurl",
        "fonds url",
        "website",
        "fund url",
    },
    "project_name": {
        "projekt",
        "projektnavn",
        "projekt navn",
        "ansoegning",
        "ansoegningstitel",
        "ansoegning titel",
        "titel",
        "project",
        "project name",
        "application",
    },
    "project_id": {
        "projekt id",
        "projektid",
        "project id",
        "project reference",
    },
    "submitted_at": {
        "dato",
        "ansoegningsdato",
        "ansoegning dato",
        "indsendelsesdato",
        "indsendt dato",
        "sendt dato",
        "soegt dato",
        "submitted",
        "submitted at",
        "application date",
    },
    "status": {
        "status",
        "resultat",
        "afgoerelse",
        "svar",
        "decision",
        "result",
    },
    "amount_requested": {
        "beloeb",
        "ansoegt beloeb",
        "ansoegningsbeloeb",
        "ansoegning beloeb",
        "soegt beloeb",
        "amount",
        "amount requested",
        "requested amount",
    },
    "external_id": {
        "id",
        "ansoegnings id",
        "ansoegningsnummer",
        "reference",
        "reference id",
        "application id",
    },
    "notes": {
        "note",
        "noter",
        "bemaerkning",
        "bemaerkninger",
        "kommentar",
        "kommentarer",
        "notes",
    },
}

_HISTORY_HEADER_LOOKUP = {
    alias: canonical
    for canonical, aliases in _HISTORY_ALIASES.items()
    for alias in aliases
}


def _history_header_map(headers: Sequence[Any]) -> dict[int, str]:
    result: dict[int, str] = {}
    used: set[str] = set()
    for index, header in enumerate(headers):
        canonical = _HISTORY_HEADER_LOOKUP.get(normalize_header(header))
        if canonical and canonical not in used:
            result[index] = canonical
            used.add(canonical)
    return result


def _header_score(headers: Sequence[Any]) -> int:
    fields = set(_history_header_map(headers).values())
    if not ({"fund_name", "fund_url"} & fields):
        return 0
    supporting = fields & {
        "project_name",
        "submitted_at",
        "status",
        "amount_requested",
        "external_id",
        "project_id",
    }
    return len(fields) + (3 if supporting else 0)


def _detect_header(rows: Sequence[Sequence[Any]], explicit_row: int | None = None) -> tuple[int, dict[int, str]]:
    if explicit_row is not None:
        if explicit_row < 1 or explicit_row > len(rows):
            raise ImportFormatError("Den angivne header-række findes ikke")
        index = explicit_row - 1
        mapping = _history_header_map(rows[index])
        if _header_score(rows[index]) == 0:
            raise ImportFormatError("Den angivne header-række har ingen genkendelig fondskolonne")
        return index, mapping

    best_index = -1
    best_score = 0
    for index, row in enumerate(rows[:25]):
        score = _header_score(row)
        if score > best_score:
            best_index = index
            best_score = score
    if best_index < 0:
        raise ImportFormatError(
            "Kunne ikke finde en header med fond/pulje eller fonds-URL"
        )
    return best_index, _history_header_map(rows[best_index])


def _history_record(
    values: Sequence[Any],
    mapping: Mapping[int, str],
    *,
    source_kind: str,
    source_name: str,
    row_number: int,
    include_notes: bool = False,
) -> dict[str, Any] | None:
    record: dict[str, Any] = {}
    for index, canonical in mapping.items():
        if index >= len(values) or values[index] in (None, ""):
            continue
        if canonical == "notes" and not include_notes:
            continue
        value = values[index]
        if canonical == "submitted_at":
            record[canonical] = normalize_date_value(value)
        else:
            record[canonical] = value
    if not record:
        return None
    if not str(record.get("fund_name", "")).strip() and not normalize_domain(
        record.get("fund_url", "")
    ):
        return None
    record["source_kind"] = source_kind
    record["source_name"] = source_name
    # Rækkenummeret er kun en flygtig locator: sortering eller indsatte rækker
    # må ikke ændre ansøgningens identitet. Kun et ægte ansøgnings-ID bruges
    # som external_id af IndexStore.
    record["extra"] = {"source_row_at_import": row_number}
    return record


def _read_csv_rows(path: Path) -> list[list[str]]:
    raw = path.read_bytes()
    text: str | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ImportFormatError("CSV-filen kunne ikke afkodes som UTF-8 eller Windows-1252")
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        delimiter = ";" if sample.count(";") >= sample.count(",") else ","
        dialect = type("FallbackDialect", (csv.excel,), {"delimiter": delimiter})
    return [list(row) for row in csv.reader(io.StringIO(text, newline=""), dialect)]


def _iter_xlsx_history_tables(
    path: Path,
    sheet_name: str | None,
    header_row: int | None,
    include_notes: bool = False,
) -> Iterator[tuple[str, int, list[dict[str, Any]]]]:
    workbook = _load_workbook(path)
    found = False
    try:
        names = [sheet_name] if sheet_name else list(workbook.sheetnames)
        for name in names:
            if name not in workbook.sheetnames:
                if sheet_name:
                    raise ImportFormatError(f"Arket findes ikke: {sheet_name}")
                continue
            sheet = workbook[name]
            rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
            if not rows:
                continue
            try:
                header_index, mapping = _detect_header(rows, header_row)
            except ImportFormatError:
                if sheet_name:
                    raise
                continue
            found = True
            records: list[dict[str, Any]] = []
            for zero_index, values in enumerate(rows[header_index + 1 :], start=header_index + 1):
                row_number = zero_index + 1
                record = _history_record(
                    values,
                    mapping,
                    source_kind="xlsx",
                    source_name=f"{path.name}#{name}",
                    row_number=row_number,
                    include_notes=include_notes,
                )
                if record is not None:
                    records.append(record)
            yield name, header_index + 1, records
    finally:
        workbook.close()
    if not found:
        raise ImportFormatError("Ingen genkendelig ansøgningshistorik fundet i XLSX-filen")


def read_application_history(
    history_path: str | Path,
    *,
    sheet_name: str | None = None,
    header_row: int | None = None,
    source_name: str | None = None,
    include_notes: bool = False,
) -> list[dict[str, Any]]:
    """Read sent-application history from flexible XLSX or CSV columns.

    Unknown columns and free-text notes are intentionally ignored by default.
    ``include_notes=True`` is an explicit private opt-in. This minimizes
    accidental ingestion of contact/member information while preserving the
    fields needed to avoid duplicate applications.
    """

    path = Path(history_path)
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        rows = _read_csv_rows(path)
        header_index, mapping = _detect_header(rows, header_row)
        records = []
        effective_source = source_name or path.name
        for zero_index, values in enumerate(rows[header_index + 1 :], start=header_index + 1):
            record = _history_record(
                values,
                mapping,
                source_kind="csv",
                source_name=effective_source,
                row_number=zero_index + 1,
                include_notes=include_notes,
            )
            if record is not None:
                records.append(record)
        return records
    if suffix in {".xlsx", ".xlsm"}:
        records: list[dict[str, Any]] = []
        for name, _header, table_records in _iter_xlsx_history_tables(
            path, sheet_name, header_row, include_notes
        ):
            effective_source = source_name or f"{path.name}#{name}"
            for record in table_records:
                record["source_name"] = effective_source
            records.extend(table_records)
        return records
    raise ImportFormatError("Historik skal være en .xlsx-, .xlsm- eller .csv-fil")


def import_application_history(
    history_path: str | Path,
    store: IndexStore,
    *,
    sheet_name: str | None = None,
    header_row: int | None = None,
    source_name: str | None = None,
    include_notes: bool = False,
) -> ImportResult:
    """Import sent applications idempotently from XLSX or CSV."""

    path = Path(history_path)
    records = read_application_history(
        path,
        sheet_name=sheet_name,
        header_row=header_row,
        source_name=source_name,
        include_notes=include_notes,
    )
    result = ImportResult(source=source_name or path.name)
    for record in records:
        result.records_seen += 1
        before = store.stats()["history"]
        try:
            store.add_history(record)
        except (TypeError, ValueError) as exc:
            result.skipped += 1
            result.errors.append(f"Historikrække {result.records_seen}: {exc}")
            continue
        after = store.stats()["history"]
        result.imported += 1
        if after > before:
            result.inserted += 1
        else:
            result.duplicates += 1
    return result


def import_sent_history(
    history_path: str | Path,
    store: IndexStore,
    **options: Any,
) -> ImportResult:
    """Concise alias for :func:`import_application_history`."""

    return import_application_history(history_path, store, **options)


def _allowed_history_host(host: str) -> bool:
    normalized = host.casefold().rstrip(".")
    return normalized in {"1drv.ms", "onedrive.live.com"} or normalized.endswith(
        ".sharepoint.com"
    )


def _validate_history_url(url: str) -> None:
    parsed = urlsplit(url)
    if parsed.scheme.casefold() != "https" or not parsed.hostname:
        raise HistoryDownloadError("Historiklinket skal være en HTTPS-URL")
    if not _allowed_history_host(parsed.hostname):
        raise HistoryDownloadError(
            "Kun 1drv.ms, onedrive.live.com og *.sharepoint.com er tilladt"
        )
    if parsed.username or parsed.password:
        raise HistoryDownloadError("URL'er med indlejrede credentials er ikke tilladt")


class _RestrictedRedirectHandler(HTTPRedirectHandler):
    def redirect_request(self, req, fp, code, msg, headers, newurl):  # type: ignore[no-untyped-def]
        _validate_history_url(newurl)
        return super().redirect_request(req, fp, code, msg, headers, newurl)


def _sanitized_source_url(url: str) -> str:
    parsed = urlsplit(url)
    return urlunsplit((parsed.scheme, parsed.netloc, parsed.path, "", ""))


def download_shared_workbook(
    url: str,
    destination_directory: str | Path,
    *,
    max_bytes: int = DEFAULT_MAX_HISTORY_DOWNLOAD_BYTES,
    opener: Any = None,
) -> Path:
    """Download an allowlisted shared XLSX with size and ZIP-content checks."""

    _validate_history_url(url)
    if max_bytes <= 0:
        raise ValueError("max_bytes must be positive")
    target_directory = Path(destination_directory)
    target_directory.mkdir(parents=True, exist_ok=True)
    target = target_directory / "shared-application-history.xlsx"
    client = opener or build_opener(_RestrictedRedirectHandler())
    request = Request(
        url,
        headers={
            "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "User-Agent": "bredballeif-fondsansoegning/1.0",
        },
    )
    try:
        response = client.open(request, timeout=30)
    except HistoryDownloadError as exc:
        raise HistoryDownloadError(
            "Delingslinket redirectede uden for de tilladte OneDrive/SharePoint-hosts. "
            "Brug en lokal eksport, connector eller et direkte XLSX-downloadlink."
        ) from exc
    except (HTTPError, URLError, TimeoutError) as exc:
        raise HistoryDownloadError(f"Kunne ikke hente den delte projektfil: {exc}") from exc

    try:
        final_url = response.geturl()
        _validate_history_url(final_url)
        content_length = response.headers.get("Content-Length")
        if content_length:
            try:
                announced_size = int(content_length)
            except ValueError:
                announced_size = 0
            if announced_size > max_bytes:
                raise HistoryDownloadError("Den delte projektfil overstiger størrelsesgrænsen")

        total = 0
        with target.open("wb") as output:
            while True:
                chunk = response.read(min(64 * 1024, max_bytes + 1 - total))
                if not chunk:
                    break
                total += len(chunk)
                if total > max_bytes:
                    raise HistoryDownloadError(
                        "Den delte projektfil overstiger størrelsesgrænsen"
                    )
                output.write(chunk)
    finally:
        response.close()

    try:
        with target.open("rb") as stream:
            magic = stream.read(4)
        if magic != b"PK\x03\x04":
            raise HistoryDownloadError(
                "Linket returnerede ikke en XLSX-fil (ofte en login-/HTML-side). "
                "Brug SharePoint-connectoren eller eksportér filen lokalt."
            )
        _validate_xlsx_archive(target)
    except (BadZipFile, ImportFormatError) as exc:
        raise HistoryDownloadError(f"Den hentede fil blev afvist som XLSX: {exc}") from exc
    return target


def import_history_from_url(
    url: str,
    store: IndexStore,
    *,
    max_bytes: int = DEFAULT_MAX_HISTORY_DOWNLOAD_BYTES,
    opener: Any = None,
    sheet_name: str | None = None,
    header_row: int | None = None,
    include_notes: bool = False,
) -> ImportResult:
    """Temporarily download and import an explicit allowlisted history URL."""

    with tempfile.TemporaryDirectory(prefix="bredballeif-fondshistorik-") as directory:
        path = download_shared_workbook(
            url, directory, max_bytes=max_bytes, opener=opener
        )
        result = import_application_history(
            path,
            store,
            sheet_name=sheet_name,
            header_row=header_row,
            source_name=_sanitized_source_url(url),
            include_notes=include_notes,
        )
    return result


__all__ = [
    "DEFAULT_MAX_HISTORY_DOWNLOAD_BYTES",
    "FUND_HEADER_ROW",
    "DGI_FUND_SHEET",
    "KNOWN_FUND_SHEETS",
    "HistoryDownloadError",
    "ImportDependencyError",
    "ImportFormatError",
    "ImportResult",
    "download_shared_workbook",
    "import_application_history",
    "import_dgi_workbook",
    "import_fund_workbook",
    "import_history_from_url",
    "import_sent_history",
    "normalize_date_value",
    "normalize_header",
    "read_application_history",
    "read_fund_workbook",
]
